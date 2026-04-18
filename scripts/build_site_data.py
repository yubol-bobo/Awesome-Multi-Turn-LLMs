from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - runtime dependency check
    raise SystemExit(
        "openpyxl is required to build the site data. Install it with `pip install openpyxl`."
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
README_PATH = ROOT / "README.md"
OUTPUT_PATH = ROOT / "docs" / "assets" / "data" / "survey-data.json"

SUMMARY_CANDIDATES = [
    ROOT / "v1_summary.xlsx",
    ROOT / "v2_summary.xlsx",
    ROOT / "overleaf" / "v1_summary.xlsx",
    ROOT / "overleaf" / "v2_summary.xlsx",
]


README_SECTION_META = {
    "Instruction Following in General (Mixed)": {
        "slug": "if-general",
        "pillar": "Tasks",
        "track": "Instruction Following",
        "cluster": "General (Mixed)",
        "group": "Instruction Following Tasks",
    },
    "Instruction Following in Math": {
        "slug": "if-math",
        "pillar": "Tasks",
        "track": "Instruction Following",
        "cluster": "Math",
        "group": "Instruction Following Tasks",
    },
    "Instruction Following in Coding": {
        "slug": "if-coding",
        "pillar": "Tasks",
        "track": "Instruction Following",
        "cluster": "Coding",
        "group": "Instruction Following Tasks",
    },
    "Instruction Following in Discussion": {
        "slug": "if-discussion",
        "pillar": "Tasks",
        "track": "Instruction Following",
        "cluster": "Discussion",
        "group": "Instruction Following Tasks",
    },
    "Conversational Engagement in General (Mixed)": {
        "slug": "ce-general",
        "pillar": "Tasks",
        "track": "Conversational Engagement",
        "cluster": "General (Mixed)",
        "group": "Conversational Engagement Tasks",
    },
    "Conversational Engagement in Roleplay": {
        "slug": "ce-roleplay",
        "pillar": "Tasks",
        "track": "Conversational Engagement",
        "cluster": "Roleplay",
        "group": "Conversational Engagement Tasks",
    },
    "Conversational Engagement in Healthcare": {
        "slug": "ce-healthcare",
        "pillar": "Tasks",
        "track": "Conversational Engagement",
        "cluster": "Healthcare",
        "group": "Conversational Engagement Tasks",
    },
    "Conversational Engagement in Education": {
        "slug": "ce-education",
        "pillar": "Tasks",
        "track": "Conversational Engagement",
        "cluster": "Education",
        "group": "Conversational Engagement Tasks",
    },
    "Conversational Engagement in Jailbreak": {
        "slug": "ce-jailbreak",
        "pillar": "Tasks",
        "track": "Conversational Engagement",
        "cluster": "Jailbreak",
        "group": "Conversational Engagement Tasks",
    },
    "In-Context Learning": {
        "slug": "impr-icl",
        "pillar": "Improvements",
        "track": "Model-Centric Approaches",
        "cluster": "In-Context Learning",
        "group": "Multi-Round Communication",
    },
    "Supervised Fine-Tuning": {
        "slug": "impr-sft",
        "pillar": "Improvements",
        "track": "Model-Centric Approaches",
        "cluster": "Supervised Fine-Tuning",
        "group": "Multi-Round Communication",
    },
    "Reinforcement Learning": {
        "slug": "impr-rl",
        "pillar": "Improvements",
        "track": "Model-Centric Approaches",
        "cluster": "Reinforcement Learning",
        "group": "Multi-Round Communication",
    },
    "New Architectures": {
        "slug": "impr-architectures",
        "pillar": "Improvements",
        "track": "Model-Centric Approaches",
        "cluster": "New Architectures",
        "group": "Multi-Round Communication",
    },
    "Single-Agent Approaches": {
        "slug": "impr-single-agent",
        "pillar": "Improvements",
        "track": "Agent-Based Approaches",
        "cluster": "Single-Agent Approaches",
        "group": "Agent-Based Approaches",
    },
    "Multi-Agent Approaches": {
        "slug": "impr-multi-agent",
        "pillar": "Improvements",
        "track": "Agent-Based Approaches",
        "cluster": "Multi-Agent Approaches",
        "group": "Agent-Based Approaches",
    },
    "Memory-Augmented Methods": {
        "slug": "impr-memory",
        "pillar": "Improvements",
        "track": "External Information Integration",
        "cluster": "Memory-Augmented Methods",
        "group": "External Information Integration",
    },
    "Retrieval-Augmented Generation": {
        "slug": "impr-rag",
        "pillar": "Improvements",
        "track": "External Information Integration",
        "cluster": "Retrieval-Augmented Generation",
        "group": "External Information Integration",
    },
    "Knowledge Graph Integration": {
        "slug": "impr-kg",
        "pillar": "Improvements",
        "track": "External Information Integration",
        "cluster": "Knowledge Graph Integration",
        "group": "External Information Integration",
    },
}


SUMMARY_SECTION_TO_SLUG = {
    "ifgeneral": "if-general",
    "ifmath": "if-math",
    "ifcoding": "if-coding",
    "ifdiscussion": "if-discussion",
    "ceoverview": "ce-general",
    "cegeneral": "ce-general",
    "ceroleplay": "ce-roleplay",
    "cehealthcare": "ce-healthcare",
    "ceeducation": "ce-education",
    "cejailbreak": "ce-jailbreak",
    "imprmcincontextlearning": "impr-icl",
    "imprmcsupervisedfinetuning": "impr-sft",
    "imprmcreinforcementlearning": "impr-rl",
    "imprmcnewarchitectures": "impr-architectures",
    "impragentsingleagentapproaches": "impr-single-agent",
    "impragentmultiagentapproaches": "impr-multi-agent",
    "impreimemoryaugmentedmethods": "impr-memory",
    "impreiretrievalaugmentedgeneration": "impr-rag",
    "impreiknowledgegraphintegration": "impr-kg",
}


TITLE_OVERRIDES_BY_KEY = {
    "chen2024roleinteract": "SocialBench: Sociality Evaluation of Role-Playing Conversational Agents",
    "yang2024aqa": "AQA-Bench: An Interactive Benchmark for Evaluating LLMs' Sequential Reasoning Ability",
    "ren2024derailyourselfmultiturnllm": "Derail Yourself: Multi-turn LLM Jailbreak Attack through Self-discovered Clues",
}


KNOWN_TAGS = ["Dataset", "Benchmark", "SFT", "RL", "Other"]


@dataclass
class ReadmeEntry:
    title: str
    venue_label: str
    venue_url: str
    links: list[dict[str, str]]
    tags: list[str]
    section_slug: str


def normalize_identifier(value: str) -> str:
    text = value.lower()
    text = text.replace("&", "and")
    return re.sub(r"[^a-z0-9]+", "", text)


def normalize_title(value: str) -> str:
    text = value.casefold()
    text = text.replace("’", "'").replace("`", "'").replace("“", '"').replace("”", '"')
    text = text.replace("&", " and ")
    text = text.replace("llms", "large language models")
    text = text.replace("llm", "large language model")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def slugify(value: str) -> str:
    text = value.casefold()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def load_summary_paths() -> list[Path]:
    seen = set()
    paths = []
    for candidate in SUMMARY_CANDIDATES:
        if candidate.exists() and candidate.name not in seen:
            paths.append(candidate)
            seen.add(candidate.name)
    if len(paths) < 2:
        raise SystemExit("Unable to find both v1_summary.xlsx and v2_summary.xlsx.")
    return paths


def load_summary_rows(paths: Iterable[Path]) -> dict[str, list[dict[str, str]]]:
    sections: dict[str, list[dict[str, str]]] = defaultdict(list)
    seen_by_section: dict[str, set[str]] = defaultdict(set)

    for path in paths:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        for raw_section, citation_key, title, authors in sheet.iter_rows(min_row=2, values_only=True):
            if not raw_section or not title:
                continue
            slug = SUMMARY_SECTION_TO_SLUG.get(normalize_identifier(str(raw_section)))
            if not slug:
                continue
            citation_key = str(citation_key or "").strip()
            authors = str(authors or "").strip()
            display_title = TITLE_OVERRIDES_BY_KEY.get(citation_key, str(title).strip())
            marker = citation_key or normalize_title(display_title)
            if marker in seen_by_section[slug]:
                continue
            sections[slug].append(
                {
                    "citationKey": citation_key,
                    "summaryTitle": str(title).strip(),
                    "displayTitle": display_title,
                    "authors": authors,
                }
            )
            seen_by_section[slug].add(marker)
    return sections


def parse_readme(readme_path: Path) -> dict[str, list[ReadmeEntry]]:
    lines = readme_path.read_text(encoding="utf-8", errors="ignore").splitlines()
    entries_by_section: dict[str, list[ReadmeEntry]] = defaultdict(list)

    current_section_slug = None
    last_entry: ReadmeEntry | None = None
    in_catalog = False

    for line in lines:
        if line.startswith("## Multi-Turn LLM Tasks"):
            in_catalog = True
            continue
        if line.startswith("## Open Challenges"):
            break
        if not in_catalog:
            continue
        if line.startswith("#### "):
            heading = line[5:].strip()
            meta = README_SECTION_META.get(heading)
            current_section_slug = meta["slug"] if meta else None
            last_entry = None
            continue
        if current_section_slug and line.startswith("- "):
            stripped = line[2:].strip()
            marker = stripped.find(" [[")
            title = stripped[:marker].strip() if marker != -1 else stripped
            link_matches = re.findall(r"\[\[([^\]]+)\]\(([^)]+)\)\]", stripped)
            if not link_matches:
                continue
            venue_label, venue_url = link_matches[0]
            extra_links = [
                {"label": label, "url": url}
                for label, url in link_matches[1:]
            ]
            entry = ReadmeEntry(
                title=title,
                venue_label=venue_label,
                venue_url=venue_url,
                links=extra_links,
                tags=[],
                section_slug=current_section_slug,
            )
            entries_by_section[current_section_slug].append(entry)
            last_entry = entry
            continue
        if last_entry and "img.shields.io/badge/" in line:
            tags = [tag for tag in KNOWN_TAGS if f"badge/{tag}" in line]
            if tags:
                last_entry.tags = tags
    return entries_by_section


def match_summary_row(readme_title: str, candidates: list[dict[str, str]]) -> tuple[int | None, dict[str, str] | None]:
    if not candidates:
        return None, None

    normalized_target = normalize_title(readme_title)

    for index, row in enumerate(candidates):
        if normalize_title(row["displayTitle"]) == normalized_target:
            return index, row

    best_index = None
    best_score = 0.0
    for index, row in enumerate(candidates):
        score = SequenceMatcher(None, normalized_target, normalize_title(row["displayTitle"])).ratio()
        if score > best_score:
            best_score = score
            best_index = index

    if best_index is not None and best_score >= 0.76:
        return best_index, candidates[best_index]
    return None, None


def build_dataset() -> dict:
    summary_paths = load_summary_paths()
    summary_rows = load_summary_rows(summary_paths)
    readme_rows = parse_readme(README_PATH)

    merged_entries = []
    unmatched_summary = {}
    unmatched_readme = []

    for heading, meta in README_SECTION_META.items():
        slug = meta["slug"]
        readme_entries = readme_rows.get(slug, [])
        available_rows = list(summary_rows.get(slug, []))

        for position, readme_entry in enumerate(readme_entries, start=1):
            match_index, summary_match = match_summary_row(readme_entry.title, available_rows)
            if summary_match is None:
                summary_match = {
                    "citationKey": "",
                    "summaryTitle": readme_entry.title,
                    "displayTitle": readme_entry.title,
                    "authors": "",
                }
                unmatched_readme.append({"section": slug, "title": readme_entry.title})
            else:
                available_rows.pop(match_index)

            merged_entries.append(
                {
                    "entryId": f"{slug}-{position:03d}",
                    "title": readme_entry.title,
                    "citationKey": summary_match["citationKey"],
                    "authors": summary_match["authors"],
                    "venue": {
                        "label": readme_entry.venue_label,
                        "url": readme_entry.venue_url,
                    },
                    "links": readme_entry.links,
                    "tags": readme_entry.tags,
                    "section": {
                        "slug": slug,
                        "title": heading,
                        "pillar": meta["pillar"],
                        "track": meta["track"],
                        "cluster": meta["cluster"],
                        "group": meta["group"],
                    },
                }
            )

        if available_rows:
            unmatched_summary[slug] = [
                {
                    "citationKey": row["citationKey"],
                    "title": row["displayTitle"],
                }
                for row in available_rows
            ]

    unique_papers: dict[str, dict] = {}
    for entry in merged_entries:
        normalized_title = normalize_title(entry["title"])
        paper_key = entry["citationKey"] or normalized_title
        if paper_key not in unique_papers:
            unique_papers[paper_key] = {
                "paperId": slugify(entry["title"]),
                "title": entry["title"],
                "citationKey": entry["citationKey"],
                "authors": entry["authors"],
                "venue": entry["venue"],
                "links": list(entry["links"]),
                "tags": list(entry["tags"]),
                "sections": [entry["section"]],
                "entryIds": [entry["entryId"]],
            }
            continue

        target = unique_papers[paper_key]
        target["entryIds"].append(entry["entryId"])
        if entry["authors"] and not target["authors"]:
            target["authors"] = entry["authors"]
        for link in entry["links"]:
            if link not in target["links"]:
                target["links"].append(link)
        for tag in entry["tags"]:
            if tag not in target["tags"]:
                target["tags"].append(tag)
        if entry["section"]["slug"] not in {section["slug"] for section in target["sections"]}:
            target["sections"].append(entry["section"])

    section_counts = Counter(entry["section"]["slug"] for entry in merged_entries)
    pillar_counts = Counter(entry["section"]["pillar"] for entry in merged_entries)
    track_counts = Counter(entry["section"]["track"] for entry in merged_entries)
    tag_counts = Counter(tag for entry in merged_entries for tag in entry["tags"])

    sections = []
    for title, meta in README_SECTION_META.items():
        sections.append(
            {
                "slug": meta["slug"],
                "title": title,
                "pillar": meta["pillar"],
                "track": meta["track"],
                "cluster": meta["cluster"],
                "group": meta["group"],
                "entryCount": section_counts.get(meta["slug"], 0),
            }
        )

    dataset = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "source": {
            "readme": str(README_PATH.relative_to(ROOT)),
            "summarySheets": [str(path.relative_to(ROOT)) for path in summary_paths],
        },
        "stats": {
            "entryCount": len(merged_entries),
            "uniquePaperCount": len(unique_papers),
            "sectionCount": len(sections),
            "pillarCounts": dict(pillar_counts),
            "trackCounts": dict(track_counts),
            "tagCounts": dict(tag_counts),
        },
        "buildDiagnostics": {
            "unmatchedReadmeCount": len(unmatched_readme),
            "unmatchedSummaryCount": sum(len(items) for items in unmatched_summary.values()),
        },
        "sections": sections,
        "entries": merged_entries,
        "papers": list(unique_papers.values()),
    }
    return dataset


def main() -> int:
    dataset = build_dataset()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(dataset, indent=2, ensure_ascii=False), encoding="utf-8")

    stats = dataset["stats"]
    print(
        json.dumps(
            {
                "output": str(OUTPUT_PATH.relative_to(ROOT)),
                "entryCount": stats["entryCount"],
                "uniquePaperCount": stats["uniquePaperCount"],
                "unmatchedReadmeCount": dataset["buildDiagnostics"]["unmatchedReadmeCount"],
                "unmatchedSummaryCount": dataset["buildDiagnostics"]["unmatchedSummaryCount"],
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
